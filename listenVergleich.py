# Listen Updater
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors

EXCEL_SPALTEN = ['A','B','C','D','E','F','G','H','I','J',
                'K','L','M','N','O','P','Q','R','S','T',
                'U','V','W','X','Y','Z','AA','AB','AC','AD',
                'AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN',
                'AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX',
                'AY','AZ']

def excelToPython(dateipfad = '_basis.xlsx', tabellenname = 'sheet',ZeilenBeginn = 1, ZeilenEnde = None):   # 1. Zile = Beschreibungen
    DATEIPFAD = dateipfad
    TABELLENNAME = tabellenname
    wb = load_workbook(filename = DATEIPFAD)
    tabelle = wb[TABELLENNAME]
    if ZeilenEnde == None:
        ZeilenEnde = int(tabelle.max_row) + 1
    print('Tabelle: "' + dateipfad + '/' + tabellenname + '" mit ' + str(ZeilenEnde) + ' Zeilen in Liste gespeichert.')

    Ergebnisliste = []                  # Ergebnis aller Zeilen
    zeilenListe = []                    # Speicherort für eine Zeilen
    spaltenListe = EXCEL_SPALTEN[0:23]  # A - X

    for i in range(ZeilenBeginn, ZeilenEnde):
        #print('____________________________________________________ ZEILE ' + str(i) + '____________________________________________________')
        zeilenListe.append(spaltenAbfragen(i, spaltenListe, tabelle))
        #print(zeilenListe)
    return zeilenListe

def spaltenAbfragen(zeile, spaltenListe, tabelle):
    spaltenInhalte = []
    for spalte in spaltenListe:
        zellenInhalt = tabelle[spalte + str(zeile)].value
        if zellenInhalt == None:
            spaltenInhalte.append("")
        else:
            spaltenInhalte.append(zellenInhalt)
    return spaltenInhalte

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')

def listeToExcel(dateipfad = None, liste = None):
    workbook = load_workbook(dateipfad, read_only = False)
    tabellenname = 'tab'
    tabelle = workbook[tabellenname]
    if dateipfad == None or liste == None or dateipfad == None and liste == None:
        print("Keine Parameter übergeben, brauche Dateipfad und Liste")
    else:
        for zeile in range(len(liste)):
            for spalte in range(len(liste[zeile])):
                zellenwert = turnNumbersIntoLetter(spalte) + str(zeile+1)
                zelle = tabelle[zellenwert]
                inhalt = liste[zeile][spalte]
                zelle.value = inhalt
                if spalte == 0 and zelle.value == "Erl":
                    tabelle[zellenwert].fill = greenFill
                elif spalte == 0 and zelle.value != "Nb":
                    tabelle[zellenwert].fill = redFill
    workbook.save(dateipfad)
    print("Exceltabelle: '"+str(dateipfad)+"' aus Liste erstellt.")

def turnNumbersIntoLetter(number):
    return EXCEL_SPALTEN[number]

def zeichenAendern(eingabe):
    zeichenVerbessertesWort = eingabe
    if "ä" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("ä","ae")
    if "ü" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("ü","ue")
    if "ö" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("ö","oe")
    if "Ä" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("Ä","Ae")
    if "Ü" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("Ü","Ue")
    if "Ö" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("Ö","Oe")
    if "ß" in eingabe:
        zeichenVerbessertesWort = eingabe.replace("ß","ss")
    return zeichenVerbessertesWort

def umlauteAendern(string):
    finalerString = ""
    for i in string:
        finalerString += (zeichenAendern(i))
    return finalerString

def vertragsnummerVergleichen(basisListe = None, vergleichsListe = None, vergleichsSpalte = 1):
    Ergebnisliste = []
    counter = 1
    zeilenzaehler = 1
    if not (basisListe == None):
        for zeile in basisListe:
            change = False
            for spalte in range(len(zeile)):
                if spalte == vergleichsSpalte: # Vertragsnummer
                    vertragsnummer_basisListe = zeile[spalte]
                    if not (vergleichsListe == None):
                        for zeile_ in vergleichsListe:
                            for spalte_ in range(len(zeile_)):
                                if spalte_ == vergleichsSpalte: # Vertragsnummer
                                    vertragsnummer_vergleichsListe = zeile_[spalte_]
                                    if vertragsnummer_basisListe == vertragsnummer_vergleichsListe and vertragsnummer_basisListe != "" and vertragsnummer_vergleichsListe != "" and zeilenzaehler != 1:
                                        Ergebnisliste.append(zeile_)
                                        print("Zeile: "+str(zeilenzaehler)+" - Änderung " + str(counter) + ' an Basisliste vorgenommen. Vergelichsattribut: ' + str(basisListe[0][vergleichsSpalte]) + " '" + str(vertragsnummer_basisListe) + "." )
                                        counter += 1
                                        change = True
                                        zeilenzaehler += 1
                    else:
                        print("Es wurde keine vergleichsListe angegeben.")
                        return
            if change == False:
                print("Zeile: "+str(zeilenzaehler)+" - Keine Änderung an Basisliste vorgenommen.")
                Ergebnisliste.append(zeile)
                zeilenzaehler += 1
    return Ergebnisliste

basis = excelToPython()
tabelle1 = excelToPython('_1.xlsx','tab')
tabelle2 = excelToPython('_2.xlsx','tab')
tabelle3 = excelToPython('_3.xlsx','tab')
tabelle4 = excelToPython('_4.xlsx','tab')
tabelle5 = excelToPython('_5.xlsx','tab')
tabelle6 = excelToPython('_6.xlsx','tab')
tabelle7 = excelToPython('_7.xlsx','tab')
tabelle8 = excelToPython('_8.xlsx','tab')
tabelle9 = excelToPython('_9.xlsx','tab')
tabelle10 = excelToPython('_10.xlsx','tab')

basis1 = vertragsnummerVergleichen(basis, tabelle1)
basis2 = vertragsnummerVergleichen(basis1, tabelle2)
basis3 = vertragsnummerVergleichen(basis2, tabelle3)
basis4 = vertragsnummerVergleichen(basis3, tabelle4)
basis5 = vertragsnummerVergleichen(basis4, tabelle5)
basis6 = vertragsnummerVergleichen(basis5, tabelle6)
basis7 = vertragsnummerVergleichen(basis6, tabelle7)
basis8 = vertragsnummerVergleichen(basis7, tabelle8)
basis9 = vertragsnummerVergleichen(basis8, tabelle9)
basis10 = vertragsnummerVergleichen(basis9, tabelle10)
listeToExcel('_test.xlsx', basis10)
